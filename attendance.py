from deepface import DeepFace
import pandas as pd
import os
import openpyxl 
from datetime import date

photos=[]
all=[]
present=[]

with os.scandir('C:\\Users\\dhruva malik\\Documents\\Project\\Present') as entries1:
    for entry1 in entries1:
        photos.append(entry1.name)
with os.scandir('C:\\Users\\dhruva malik\\Documents\\Project\\All') as entries2:
    for entry2 in entries2:
        all.append(entry2.name)
i=0
j=0
while i<len(photos):
    j=0
    d=len(photos)
    while j<len(all):
        result=DeepFace.verify("C:\\Users\\dhruva malik\\Documents\\Project\\Present\\"+photos[i],"C:\\Users\\dhruva malik\\Documents\\Project\\All\\"+all[j])
        
        if result["verified"]==True:
            present.append(all[j])
            del all[j]
            del photos[i]
            break
        j+=1
    if len(photos)==d:
        del photos[i]
present_names=[]
for i in range(len(present)):
    present_names.append(present[i][0:-4])

path = "Book1.xlsx"
wb_obj = openpyxl.load_workbook(path)
wb=wb_obj['Sheet1']
sheet_obj = wb_obj.active
column1=sheet_obj.max_column
row1=sheet_obj.max_row
current_column=column1+1
wb.cell(1,current_column).value=str(date.today())
for j in range(0,len(present_names)):
    for i in range(1,row1+1):
        cell_obj = sheet_obj.cell(row = i, column = 2) 
        if cell_obj.value==present_names[j]:
            wb.cell(i,current_column).value="Present"
for k in range(1,row1+1):
#    print(wb.cell(i,current_column).value)
    if sheet_obj.cell(row = k, column = current_column).value==None:
        wb.cell(k,current_column).value='Absent'
        
wb_obj.save("Book1.xlsx")

from os import listdir
Target = "C:\\Users\\dhruva malik\\Documents\\Project\\Present\\"
for X in listdir("C:\\Users\\dhruva malik\\Documents\\Project\\Present"):
    if X.endswith('.jpg'):
        os.remove( Target + X)
print("Attendance has been successfully marked")        
        
